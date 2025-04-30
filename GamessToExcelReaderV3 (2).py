from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === FUNCTIONS ===

def extract_bond_length(text: str) -> float | None:
    """Extract the first bond length from the INTERNUCLEAR DISTANCES section."""
    match = re.search(
        r"INTERNUCLEAR DISTANCES \(ANGS\.\).*?\n\s*\d+\s+[A-Z]+\s+([0-9]+\.\d+)\s+\*",
        text,
        re.DOTALL
    )
    return float(match.group(1)) if match else None

def extract_heat_of_formation(text: str) -> float | None:
    """Extract the heat of formation from the DENSITY CONVERGED section."""
    match = re.search(r"HEAT OF FORMATION IS\s+(-?\d+\.\d+)", text)
    return float(match.group(1)) if match else None

def extract_total_energy(text:str) -> float |None:
    """Extract the total energy from the ENERGY COMPONENTS section"""
    match = re.search(r'TOTAL ENERGY\s+=\s+(-?\d+\.\d+)',text)
    return float(match.group(1)) if match else None

def parse_filename(file: Path):
    """Split filename into molecule, force field, basis set, and computational method."""
    try:
        parts = file.stem.split('_')
        molecule, forcefield, basis, comp = parts
        return molecule, forcefield, basis, comp
    except ValueError:
        return None, None, None, None

def run_time(text:str)-> str | None:
    matches = re.findall(r'TOTAL WALL CLOCK TIME=\s+([0-9]+\.\d+)',text)
    return matches[-1] if matches else None

def fallback_parse_input_section(text: str):
    """Extract molecule, comp method, and basis from the input section."""
    try:
        match = re.search(r"INPUT CARD>!\s*(.*?)\s*\|\s*.*?\|\s*(.*?)\s*$", text, re.MULTILINE)
        if not match:
            return None, None, None, None
        molecule = match.group(1).strip()
        method_basis = match.group(2).strip()
        if '/' in method_basis:
            method, basis = method_basis.split('/', 1)
        else:
            method, basis = method_basis, 'Unknown'
        return molecule, 'Unknown', basis.strip(), method.strip()
    except Exception:
        return None, None, None, None


# === PATHS ===

input_path = Path(r'C:\Users\Public\gamess-64\outputs')
output_path = Path(r'C:\Users\Public\gamess-64\saved outputs\Python compiled csv')
output_path.mkdir(parents=True, exist_ok=True)
excel_file = output_path / 'Gamess_summary.xlsx'



# === MAIN LOOP ===

results = []

for file in input_path.glob('*.log'):
    if file.name.lower() == 'readme.txt':
        print(f"Skipping {file.name}")
        continue
    
    try:
        
        
        with file.open(encoding='utf-8', errors='ignore') as f:
            contents = f.read()
        #fallback if you didn't name the files like me
        molecule, ff, basis, method = parse_filename(file)
        if None in (molecule,ff,basis, method):
            molecule,ff,basis,method = fallback_parse_input_section(contents)
        
        bond = extract_bond_length(contents)
        heat = extract_heat_of_formation(contents)
        energy = extract_total_energy(contents)        
        time = run_time(contents)
        #check if failed
        if "EXECUTION OF GAMESS TERMINATED NORMALLY" not in contents:
            results.append({
               "molecule": molecule,
                "force_field": ff,
                "basis": basis,
                "comp_method": method,
                "bond_length (Å)": bond if bond is not None else 'NA',
                "heat_of_formation (kcal/mol)": heat if heat is not None else 'NA',
                "Total Energy (Hartrees)": energy if energy is not None else "NA",
                "Run_Time (s)": time if time is not None else "NA",
                "Completion": (f"Incomplete")
                })
            continue
       
        

        #else we continue to a succesful compile
        if any(v is not None for v in [bond,heat,energy]):
            results.append({
                "molecule": molecule,
                "force_field": ff,
                "basis": basis,
                "comp_method": method,
                "bond_length (Å)": bond if bond is not None else 'NA',
                "heat_of_formation (kcal/mol)": heat if heat is not None else 'NA',
                "Total Energy (Hartrees)": energy if energy is not None else "NA",
                "Run_Time (s)": time if time is not None else "NA",
                "Completion": (f"Completed")
            })
        
        
        else:
            print(f'Failed Entirely')
    except Exception as e:
        print (f'error processing{file.name}')
# === EXPORT ===
df = pd.DataFrame(results)
df.to_excel(excel_file, index = False)

wb = load_workbook(excel_file)
ws = wb.active

#THis is all to make the "completion" results either red or green
#df.to_csv(output_file, index=False, encoding='utf-8-sig')
completion_col = None
for idx, cell in enumerate(ws[1], 1):  # row 1, 1-based indexing
    if cell.value == "Completion":
                completion_col = idx
                break

        # Define fills
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Apply fill based on value
for row in range(2, ws.max_row + 1):  # skip header row
    cell = ws.cell(row=row, column=completion_col)
    if cell.value == "Completed":
        cell.fill = green_fill
    elif cell.value == "Incomplete":
        cell.fill = red_fill



wb.save(excel_file)

print(f"\n✅ XLSX saved to:\n{excel_file}")
